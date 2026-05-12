# app/utils/legacy/excel_filter_ai.py
# -*- coding: utf-8 -*-
"""
EXCEL FILTER AI – Logic lọc dữ liệu, phát hiện lỗi lớp, ánh xạ học sinh.
ĐÃ LOẠI BỎ PHẦN GIAO DIỆN TKINTER.
"""

import pandas as pd
import numpy as np
import re
import os
import json
import logging
from datetime import datetime
from typing import Tuple, List, Dict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import Counter

# ============================================================================
# CẤU HÌNH
# ============================================================================
class Config:
    APP_NAME = "EXCEL FILTER AI"
    VERSION = "8.0"
    DEFAULT_ENCODING = "utf-8"
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls", ".xlsm"]
    MAX_HISTORY_SIZE = 10
    DEFAULT_PREVIEW_ROWS = 15
    LOG_FILE = "excel_filter.log"
    STUDENT_LIST_FILE = "student_list.json"
    APP_DATA_DIR = "app_data"

# ============================================================================
# LOGGING
# ============================================================================
def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[logging.StreamHandler()]
    )
    return logging.getLogger(__name__)

# ============================================================================
# LỚP XỬ LÝ CHÍNH
# ============================================================================
class ExcelFilterApp:
    """Lọc file Excel, phân tích AI (rule-based), quản lý danh sách học sinh."""
    def __init__(self):
        self.logger = setup_logging()
        self.original_df = None
        self.filtered_df = None
        self.split_all_students = True
        self.file_path = None

        self.app_data_dir = Config.APP_DATA_DIR
        os.makedirs(self.app_data_dir, exist_ok=True)

        self.student_list_path = os.path.join(self.app_data_dir, Config.STUDENT_LIST_FILE)
        self.student_mapping = {}
        self._load_student_list_from_file()

        self.history = []
        self.current_index = -1
        self.progress_callback = None
        self.status_callback = None

        self.pattern = r'([^,(]+?)\s*\(\s*(.*?)\s*\)(?=\s*,\s*|$)'   # pattern cho "tên (nhận xét)"
        self.negative_keywords = [
            "ồn", "chưa nghiêm túc", "không nghiêm túc", "thiếu nghiêm túc",
            "chưa hoàn thành", "không hoàn thành", "không tham gia",
            "quên", "không mang", "không ghi", "chưa thuộc bài", "không thuộc bài",
            "nói chuyện", "nói leo", "tự ý", "di chuyển", "chơi", "bỏ quên",
            "cần cải thiện", "cần chú ý", "cần nghiêm túc", "hạn chế",
            "không tập trung", "đạt", "được", "chưa làm", "chưa bổ sung",
            "không đúng", "chậm trễ", "vào trễ", "không đầy đủ", "còn ồn",
            "làm ồn", "mất trật tự", "không chuẩn bị", "không có", "chưa có",
            "chưa xong", "không xong", "chưa sẵn sàng", "không sẵn sàng"
        ]
        self.error_mapping = {
            "ồn": "Ồn trong lớp",
            "chưa nghiêm túc": "Không nghiêm túc",
            "không nghiêm túc": "Không nghiêm túc",
            "thiếu nghiêm túc": "Không nghiêm túc",
            "chưa hoàn thành": "Chưa hoàn thành nhiệm vụ",
            "không hoàn thành": "Chưa hoàn thành nhiệm vụ",
            "không tham gia": "Không tham gia hoạt động",
            "quên": "Quên dụng cụ/sách vở",
            "không mang": "Không mang đầy đủ dụng cụ",
            "không ghi": "Không ghi bài đầy đủ",
            "chưa thuộc bài": "Chưa thuộc bài",
            "không thuộc bài": "Chưa thuộc bài",
            "nói chuyện": "Nói chuyện riêng",
            "nói leo": "Nói leo",
            "tự ý": "Tự ý di chuyển/vi phạm nội quy",
            "di chuyển": "Tự ý di chuyển",
            "chơi": "Chơi trong giờ học",
            "bỏ quên": "Bỏ quên đồ dùng",
            "cần cải thiện": "Cần cải thiện ý thức",
            "cần chú ý": "Cần tập trung chú ý hơn",
            "cần nghiêm túc": "Cần nghiêm túc hơn",
            "hạn chế": "Có điểm hạn chế",
            "không tập trung": "Không tập trung",
            "đạt": "Chỉ đạt mức trung bình",
            "được": "Chỉ ở mức chấp nhận được",
            "chưa làm": "Chưa làm bài tập",
            "chưa bổ sung": "Chưa bổ sung bài đầy đủ",
            "không đúng": "Không thực hiện đúng yêu cầu",
            "chậm trễ": "Đến lớp trễ",
            "vào trễ": "Vào lớp trễ",
            "không đầy đủ": "Chuẩn bị không đầy đủ",
            "còn ồn": "Ồn trong lớp",
            "làm ồn": "Ồn trong lớp",
            "mất trật tự": "Mất trật tự",
            "không chuẩn bị": "Không chuẩn bị bài",
            "không có": "Thiếu dụng cụ/bài tập",
            "chưa có": "Thiếu dụng cụ/bài tập",
            "chưa xong": "Chưa hoàn thành nhiệm vụ",
            "không xong": "Chưa hoàn thành nhiệm vụ",
            "chưa sẵn sàng": "Chưa chuẩn bị sẵn sàng",
            "không sẵn sàng": "Chưa chuẩn bị sẵn sàng"
        }
        self.logger.info("Khởi tạo ExcelFilterApp")

    # ========== QUẢN LÝ DANH SÁCH HỌC SINH ==========
    def _load_student_list_from_file(self):
        try:
            if os.path.exists(self.student_list_path):
                with open(self.student_list_path, 'r', encoding=Config.DEFAULT_ENCODING) as f:
                    self.student_mapping = json.load(f)
                self.logger.info(f"Đã tải danh sách {len(self.student_mapping)} học sinh từ file")
                return True
        except Exception as e:
            self.logger.error(f"Lỗi khi tải danh sách học sinh: {str(e)}")
        return False

    def _save_student_list_to_file(self):
        try:
            with open(self.student_list_path, 'w', encoding=Config.DEFAULT_ENCODING) as f:
                json.dump(self.student_mapping, f, ensure_ascii=False, indent=2)
            self.logger.info(f"Đã lưu danh sách {len(self.student_mapping)} học sinh vào file")
            return True
        except Exception as e:
            self.logger.error(f"Lỗi khi lưu danh sách học sinh: {str(e)}")
            return False

    def get_student_name(self, student_id: str) -> str:
        if not student_id or pd.isna(student_id):
            return ""
        student_id_str = str(student_id).strip()
        if student_id_str in self.student_mapping:
            return self.student_mapping[student_id_str]
        # Tìm kiếm linh hoạt
        student_id_lower = student_id_str.lower()
        for key, value in self.student_mapping.items():
            if key.lower() == student_id_lower:
                return value
        clean_id = re.sub(r'[^a-zA-Z0-9]', '', student_id_str)
        clean_id_lower = clean_id.lower()
        for key, value in self.student_mapping.items():
            clean_key = re.sub(r'[^a-zA-Z0-9]', '', key)
            if clean_key.lower() == clean_id_lower:
                return value
        if '@' in student_id_str:
            username = student_id_str.split('@')[0].strip()
            if username in self.student_mapping:
                return self.student_mapping[username]
            username_lower = username.lower()
            for key, value in self.student_mapping.items():
                if key.lower() == username_lower:
                    return value
        for key, value in self.student_mapping.items():
            if student_id_str in key or key in student_id_str:
                return value
        return ""

    def add_student_manually(self, email: str, name: str) -> Tuple[bool, str]:
        if not email or not name:
            return False, "Email và tên không được để trống"
        email = email.strip()
        name = name.strip()
        self.student_mapping[email] = name
        if '@' in email:
            username = email.split('@')[0].strip()
            if username and username not in self.student_mapping:
                self.student_mapping[username] = name
        self._save_student_list_to_file()
        return True, f"Đã thêm học sinh: {email} -> {name}"

    def remove_student(self, email: str) -> Tuple[bool, str]:
        if email not in self.student_mapping:
            return False, f"Không tìm thấy học sinh: {email}"
        target_name = self.student_mapping[email]
        keys_to_remove = [k for k, v in self.student_mapping.items() if v == target_name]
        for k in keys_to_remove:
            del self.student_mapping[k]
        self._save_student_list_to_file()
        return True, f"Đã xóa học sinh: {email} -> {target_name}"

    def clear_student_list(self) -> Tuple[bool, str]:
        if not self.student_mapping:
            return False, "Danh sách học sinh đã trống"
        count = len(self.student_mapping)
        self.student_mapping.clear()
        self._save_student_list_to_file()
        return True, f"Đã xóa toàn bộ {count} học sinh"

    def load_student_list_from_excel(self, file_path: str) -> Tuple[bool, str]:
        try:
            student_df = pd.read_excel(file_path, dtype=object)
            email_col = name_col = None
            for col in student_df.columns:
                col_lower = str(col).lower()
                if any(kw in col_lower for kw in ['email', 'tài khoản', 'username', 'mã', 'id', 'tên đăng nhập']):
                    email_col = col
                elif any(kw in col_lower for kw in ['họ tên', 'tên', 'full name', 'fullname', 'họ và tên']):
                    name_col = col
            if not email_col or not name_col:
                if len(student_df.columns) >= 2:
                    email_col = student_df.columns[0]
                    name_col = student_df.columns[1]
                else:
                    return False, "File danh sách phải có ít nhất 2 cột"
            new_mapping = {}
            mapped_count = 0
            for idx, row in student_df.iterrows():
                email_val = str(row[email_col]).strip() if pd.notna(row[email_col]) else ""
                name_val = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
                if email_val and name_val:
                    new_mapping[email_val] = name_val
                    if '@' in email_val:
                        username = email_val.split('@')[0].strip()
                        if username:
                            new_mapping[username] = name_val
                    mapped_count += 1
            self.student_mapping.update(new_mapping)
            self._save_student_list_to_file()
            return True, f"Đã tải {mapped_count} học sinh từ file Excel"
        except Exception as e:
            return False, f"Lỗi khi đọc file danh sách: {str(e)}"

    # ========== ĐỌC & CHUẨN HÓA DỮ LIỆU ==========
    def validate_file(self, file_path: str) -> Tuple[bool, str]:
        if not os.path.exists(file_path):
            return False, f"File không tồn tại: {file_path}"
        if not any(file_path.lower().endswith(ext) for ext in Config.SUPPORTED_EXTENSIONS):
            return False, f"File không phải định dạng Excel hỗ trợ: {', '.join(Config.SUPPORTED_EXTENSIONS)}"
        return True, "File hợp lệ"

    def load_file(self, file_path: str) -> Tuple[bool, str]:
        self.file_path = file_path
        valid, msg = self.validate_file(file_path)
        if not valid:
            return False, msg
        try:
            excel_file = pd.ExcelFile(file_path, engine=None)
            self.original_df = pd.read_excel(excel_file, sheet_name=0, header=1, dtype=object)
            success, msg = self._standardize_columns()
            if not success:
                return False, msg
            success, msg = self.validate_data()
            if not success:
                return False, msg
            self._add_class_errors_column()
            if self.split_all_students:
                self._split_student_evaluation_expanded()
            else:
                self._split_student_evaluation()
            if self.student_mapping:
                self._add_student_names()
            self._save_state()
            return True, f"Đã tải {len(self.original_df)} dòng"
        except Exception as e:
            return False, f"Lỗi đọc file: {str(e)}"

    def _standardize_columns(self) -> Tuple[bool, str]:
        self.original_df.columns = self.original_df.columns.str.strip()
        column_mapping = {}
        for col in self.original_df.columns:
            lower_col = str(col).lower().strip()
            if 'điểm' in lower_col:
                column_mapping[col] = 'Điểm'
            elif 'nhận xét lớp' in lower_col:
                column_mapping[col] = 'Nhận xét lớp'
            elif 'đánh giá học sinh' in lower_col:
                column_mapping[col] = 'Đánh giá học sinh'
            elif 'lớp' in lower_col:
                column_mapping[col] = 'Lớp'
            elif 'môn' in lower_col:
                column_mapping[col] = 'Môn học'
            elif 'ngày' in lower_col:
                column_mapping[col] = 'Ngày'
            elif 'tiết' in lower_col:
                column_mapping[col] = 'Tiết'
            elif 'stt' in lower_col:
                column_mapping[col] = 'STT'
            elif 'người dạy' in lower_col or 'giáo viên' in lower_col:
                column_mapping[col] = 'Người dạy'
            elif 'không phép' in lower_col or 'vắng không phép' in lower_col:
                column_mapping[col] = 'Không phép'
        self.original_df.rename(columns=column_mapping, inplace=True)
        # Chuẩn hóa Điểm
        if 'Điểm' in self.original_df.columns:
            self.original_df['Điểm'] = pd.to_numeric(
                self.original_df['Điểm'].astype(str).str.replace(',', '.', regex=False), errors='coerce'
            )
            self.original_df['Điểm'] = self.original_df['Điểm'].round(1)
            self.original_df['Điểm'] = self.original_df['Điểm'].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "")
            mean_val = pd.to_numeric(self.original_df['Điểm'], errors='coerce').mean()
            if not pd.isna(mean_val):
                self.original_df['Điểm'] = self.original_df['Điểm'].apply(
                    lambda x: f"{mean_val:.1f}" if x == "" or pd.isna(x) else x
                )
        # Chuẩn hóa Ngày
        if 'Ngày' in self.original_df.columns:
            try:
                self.original_df['Ngày'] = pd.to_datetime(self.original_df['Ngày'], errors='coerce', dayfirst=True)
                self.original_df['Ngày'] = self.original_df['Ngày'].dt.strftime('%d-%m-%Y')
            except:
                pass
        # Xóa các cột không cần thiết
        columns_to_drop = ['Session', 'Style', 'Username', 'Phòng', 'Dạy AI', 'tên bài dạy', 'có phép', 'Tên bài dạy']
        for col in columns_to_drop:
            if col in self.original_df.columns:
                self.original_df.drop(columns=[col], inplace=True)
        return True, "Chuẩn hóa dữ liệu thành công"

    def validate_data(self) -> Tuple[bool, str]:
        required_columns = ['Điểm', 'Nhận xét lớp']
        missing = [col for col in required_columns if col not in self.original_df.columns]
        if missing:
            return False, f"Thiếu các cột bắt buộc: {', '.join(missing)}"
        if 'Đánh giá học sinh' not in self.original_df.columns:
            self.original_df['Đánh giá học sinh'] = ''
        return True, "Dữ liệu hợp lệ"

    # ========== PHÂN TÍCH LỖI LỚP ==========
    def detect_class_errors(self, comment):
        if pd.isna(comment) or comment is None or str(comment).strip() == "":
            return ""
        comment_lower = str(comment).lower().strip()
        errors = set()
        for keyword in self.negative_keywords:
            pattern = r'\b' + re.escape(keyword) + r'\b'
            if re.search(pattern, comment_lower):
                errors.add(self.error_mapping.get(keyword, keyword.capitalize()))
        if errors:
            return ", ".join(sorted(errors))
        return ""

    def _add_class_errors_column(self):
        if 'Nhận xét lớp' not in self.original_df.columns:
            return
        self.original_df['Lỗi của lớp'] = self.original_df['Nhận xét lớp'].apply(self.detect_class_errors)

    # ========== TÁCH ĐÁNH GIÁ HỌC SINH ==========
    def _split_student_evaluation_expanded(self):
        if 'Đánh giá học sinh' not in self.original_df.columns:
            return
        new_rows = []
        for idx, row in self.original_df.iterrows():
            eval_text = row['Đánh giá học sinh']
            if pd.notna(eval_text) and isinstance(eval_text, str) and eval_text.strip():
                matches = re.findall(self.pattern, eval_text)
                if matches:
                    for username, comment in matches:
                        new_row = row.copy()
                        new_row['học sinh'] = username.strip()
                        new_row['đánh giá'] = comment.strip()
                        if self.student_mapping:
                            student_name = self.get_student_name(username.strip())
                            new_row['Họ tên học sinh'] = student_name
                        ai_result = self._ai_classify_evaluation(comment.strip())
                        new_row['phân loại'] = ai_result['loại']
                        new_rows.append(new_row)
                else:
                    # Fallback: tách bằng dấu phẩy
                    students = self._extract_students_alternative(eval_text)
                    if students:
                        for username, comment in students:
                            new_row = row.copy()
                            new_row['học sinh'] = username.strip()
                            new_row['đánh giá'] = comment.strip()
                            if self.student_mapping:
                                student_name = self.get_student_name(username.strip())
                                new_row['Họ tên học sinh'] = student_name
                            ai_result = self._ai_classify_evaluation(comment.strip())
                            new_row['phân loại'] = ai_result['loại']
                            new_rows.append(new_row)
                    else:
                        row['học sinh'] = None
                        row['đánh giá'] = None
                        if 'Họ tên học sinh' not in row:
                            row['Họ tên học sinh'] = None
                        new_rows.append(row)
            else:
                row['học sinh'] = None
                row['đánh giá'] = None
                if 'Họ tên học sinh' not in row:
                    row['Họ tên học sinh'] = None
                new_rows.append(row)
        self.original_df = pd.DataFrame(new_rows).reset_index(drop=True)

    def _extract_students_alternative(self, text: str) -> List[Tuple[str, str]]:
        students = []
        parts = [p.strip() for p in text.split(',') if p.strip()]
        for part in parts:
            if '(' in part and ')' in part:
                name_start = part.find('(')
                name_end = part.find(')')
                if name_start < name_end:
                    username = part[:name_start].strip()
                    comment = part[name_start+1:name_end].strip()
                    students.append((username, comment))
            else:
                students.append((part, ''))
        return students

    def _split_student_evaluation(self):
        if 'Đánh giá học sinh' not in self.original_df.columns:
            return
        self.original_df['học sinh'] = None
        self.original_df['đánh giá'] = None
        self.original_df['Họ tên học sinh'] = None
        for idx, row in self.original_df.iterrows():
            eval_text = row['Đánh giá học sinh']
            if pd.notna(eval_text) and isinstance(eval_text, str) and eval_text.strip():
                matches = re.findall(self.pattern, eval_text)
                if matches:
                    first_match = matches[0]
                    username = first_match[0].strip()
                    comment = first_match[1].strip()
                    self.original_df.at[idx, 'học sinh'] = username
                    self.original_df.at[idx, 'đánh giá'] = comment
                    if self.student_mapping:
                        self.original_df.at[idx, 'Họ tên học sinh'] = self.get_student_name(username)
                    ai_result = self._ai_classify_evaluation(comment.strip())
                    self.original_df.at[idx, 'phân loại'] = ai_result['loại']

    def _ai_classify_evaluation(self, comment: str) -> dict:
        if not comment or not isinstance(comment, str):
            return {'loại': 'Không xác định'}
        comment_lower = comment.lower().strip()
        comment_lower = comment_lower.replace('chura', 'chưa').replace('btvn', 'bài tập về nhà').replace('ko', 'không').replace('k', 'không')
        categories = {
            'học_tốt': ['học tốt', 'tích cực', 'phát biểu', 'tiến bộ', 'cố gắng',
                       'hoàn thành tốt', 'xuất sắc', 'giỏi', 'nhanh', 'tốt'],
            'nói_chuyện': ['nói chuyện', 'trò chuyện', 'mất trật tự', 'ồn',
                          'quay bài', 'quay cóp', 'trao đổi', 'bàn tán'],
            'không_ghi_bài': ['chưa ghi bài', 'thiếu vở', 'không ghi', 'chép bài chậm',
                             'ghi bài thiếu', 'vở trống', 'không có vở'],
            'không_làm_bài': ['chưa làm bài', 'không làm bài', 'thiếu bài tập',
                             'chưa hoàn thành', 'bài tập chưa xong', 'btvn chưa làm'],
            'không_nghiêm_túc': ['chưa nghiêm túc', 'không nghiêm túc', 'lơ là',
                                'thiếu tập trung', 'mất tập trung', 'lười học',
                                'không chú ý', 'làm việc riêng']
        }
        if any(keyword in comment_lower for keyword in categories['học_tốt']):
            return {'loại': 'Học tốt'}
        elif any(keyword in comment_lower for keyword in categories['không_làm_bài']):
            return {'loại': 'Không làm bài'}
        elif any(keyword in comment_lower for keyword in categories['không_ghi_bài']):
            return {'loại': 'Không ghi bài'}
        elif any(keyword in comment_lower for keyword in categories['nói_chuyện']):
            return {'loại': 'Hay nói chuyện'}
        elif any(keyword in comment_lower for keyword in categories['không_nghiêm_túc']):
            return {'loại': 'Không nghiêm túc'}
        return {'loại': 'Bình thường'}

    # ========== THÊM TÊN HỌC SINH ==========
    def _add_student_names(self):
        if 'học sinh' not in self.original_df.columns:
            return
        self.original_df['Họ tên học sinh'] = ""
        for idx, row in self.original_df.iterrows():
            student_id = row.get('học sinh')
            if pd.notna(student_id) and str(student_id).strip():
                name = self.get_student_name(str(student_id))
                if name:
                    self.original_df.at[idx, 'Họ tên học sinh'] = name

    # ========== LỌC & SẮP XẾP ==========
    def apply_filters(self) -> Tuple[bool, str]:
        if self.original_df is None or len(self.original_df) == 0:
            return False, "Chưa tải dữ liệu!"
        try:
            self.filtered_df = self.original_df.copy()
            initial_count = len(self.filtered_df)
            # Định dạng văn bản
            if 'đánh giá' in self.filtered_df.columns:
                self.filtered_df['đánh giá'] = self.filtered_df['đánh giá'].apply(self._capitalize_first_word)
            if 'Lỗi của lớp' in self.filtered_df.columns:
                self.filtered_df['Lỗi của lớp'] = self.filtered_df['Lỗi của lớp'].apply(self._capitalize_first_word)
            # Sắp xếp cột
            desired_order = ['STT', 'Ngày', 'Tiết', 'Lớp', 'Môn học', 'Người dạy', 'Không phép',
                            'Đánh giá học sinh', 'học sinh', 'Họ tên học sinh', 'đánh giá',
                            'Nhận xét lớp', 'Lỗi của lớp', 'phân loại', 'Điểm']
            existing_order = [col for col in desired_order if col in self.filtered_df.columns]
            remaining = [col for col in self.filtered_df.columns if col not in existing_order]
            self.filtered_df = self.filtered_df[existing_order + remaining]
            self._save_state()
            return True, f"Đã áp dụng bộ lọc, giữ nguyên {len(self.filtered_df)} dòng"
        except Exception as e:
            return False, f"Lỗi khi lọc: {str(e)}"

    def _capitalize_first_word(self, text):
        if pd.isna(text) or text is None or str(text).strip() == "":
            return text
        text_str = str(text).strip()
        if len(text_str) > 0:
            return text_str[0].upper() + text_str[1:]
        return text_str

    # ========== LƯU FILE KẾT QUẢ ==========
    def save_filtered_file(self, output_path: str) -> Tuple[bool, str]:
        if self.filtered_df is None or len(self.filtered_df) == 0:
            return False, "Không có dữ liệu để lưu!"
        try:
            export_df = self.filtered_df.copy()
            # Gom nhóm để bớt trùng lặp "Nhận xét lớp"
            group_cols = ['STT', 'Ngày', 'Tiết', 'Lớp', 'Môn học', 'Người dạy']
            existing_group = [c for c in group_cols if c in export_df.columns]
            if 'Nhận xét lớp' in export_df.columns and existing_group:
                export_df['_group_key'] = ''
                for col in existing_group:
                    export_df['_group_key'] += export_df[col].astype(str) + '|'
                last_group = None
                for idx in range(len(export_df)):
                    current_group = export_df.at[idx, '_group_key']
                    if current_group != last_group:
                        last_group = current_group
                    else:
                        export_df.at[idx, 'Nhận xét lớp'] = ''
                export_df.drop(columns=['_group_key'], inplace=True)
            # Xóa cột thừa
            columns_to_drop = ['học tốt', 'nói chuyện', 'không ghi bài', 'không làm bài', 'không nghiêm túc']
            for col in columns_to_drop:
                if col in export_df.columns:
                    export_df.drop(columns=[col], inplace=True)
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            export_df.to_excel(output_path, index=False)
            # Định dạng bằng openpyxl
            wb = load_workbook(output_path)
            ws = wb.active
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)
            header_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                   top=Side(style="thin"), bottom=Side(style="thin"))
            data_alignment = Alignment(wrap_text=True, vertical="top")
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))
            # Định dạng header
            for col in range(1, ws.max_column+1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = header_border
            # Định dạng dữ liệu
            for row in range(2, ws.max_row+1):
                for col in range(1, ws.max_column+1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = data_alignment
                    cell.border = thin_border
            # Tự động độ rộng cột
            for col in range(1, ws.max_column+1):
                max_length = 0
                col_letter = get_column_letter(col)
                for row in range(1, ws.max_row+1):
                    val = ws.cell(row=row, column=col).value
                    if val:
                        max_length = max(max_length, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            wb.save(output_path)
            return True, f"Đã lưu file: {output_path}"
        except Exception as e:
            return False, f"Lỗi khi lưu file: {str(e)}"

    # ========== PHÂN TÍCH THỐNG KÊ ==========
    def analyze_students_ai(self) -> Tuple[bool, str]:
        if self.filtered_df is None or len(self.filtered_df) == 0:
            return False, "Không có dữ liệu!"
        try:
            for idx, row in self.filtered_df.iterrows():
                if pd.notna(row.get('đánh giá')):
                    ai_result = self._ai_classify_evaluation(str(row['đánh giá']))
                    self.filtered_df.at[idx, 'phân loại'] = ai_result['loại']
            stats = self.filtered_df['phân loại'].value_counts()
            total = len(self.filtered_df)
            msg = f"Phân loại ({total} học sinh):\n" + "\n".join(f"- {k}: {v}" for k, v in stats.items())
            return True, msg
        except Exception as e:
            return False, f"Lỗi phân tích AI: {str(e)}"

    def analyze_class_errors(self) -> Tuple[bool, str]:
        if self.filtered_df is None or len(self.filtered_df) == 0:
            return False, "Không có dữ liệu!"
        classes_with_errors = self.filtered_df[self.filtered_df['Lỗi của lớp'] != '']
        if len(classes_with_errors) == 0:
            return False, "Không có lớp nào có lỗi"
        all_errors = []
        for errors in classes_with_errors['Lỗi của lớp']:
            all_errors.extend(errors.split(", "))
        error_counts = Counter(all_errors)
        msg = "Lớp có lỗi:\n" + "\n".join(f"- {k}: {v} lớp" for k, v in error_counts.most_common(10))
        return True, msg

    def filter_classes_with_errors(self) -> Tuple[bool, str]:
        if self.filtered_df is None or len(self.filtered_df) == 0:
            return False, "Không có dữ liệu!"
        self.filtered_df = self.filtered_df[self.filtered_df['Lỗi của lớp'] != '']
        return True, f"Đã lọc, còn {len(self.filtered_df)} lớp có lỗi"

    def get_preview(self, num_rows: int = 10):
        if self.filtered_df is not None and len(self.filtered_df) > 0:
            return self.filtered_df.head(num_rows)
        elif self.original_df is not None and len(self.original_df) > 0:
            return self.original_df.head(num_rows)
        else:
            return pd.DataFrame()

    def get_student_list_info(self) -> str:
        if not self.student_mapping:
            return "Chưa có danh sách học sinh"
        name_to_emails = {}
        for email, name in self.student_mapping.items():
            name_to_emails.setdefault(name, []).append(email)
        total = len(name_to_emails)
        return f"{total} học sinh, {len(self.student_mapping)} email/username"

    # ========== UNDO / REDO ==========
    def _save_state(self):
        if self.original_df is None or len(self.original_df) == 0:
            return
        state = {
            'original_df': self.original_df.copy(),
            'filtered_df': self.filtered_df.copy() if self.filtered_df is not None else None,
            'split_all_students': self.split_all_students,
            'student_mapping': self.student_mapping.copy() if self.student_mapping else {},
            'timestamp': datetime.now()
        }
        if len(self.history) >= Config.MAX_HISTORY_SIZE:
            self.history.pop(0)
        self.history.append(state)
        self.current_index = len(self.history) - 1

    def undo(self) -> Tuple[bool, str]:
        if self.current_index <= 0:
            return False, "Không thể hoàn tác thêm"
        self.current_index -= 1
        state = self.history[self.current_index]
        self.original_df = state['original_df'].copy()
        self.filtered_df = state['filtered_df'].copy() if state['filtered_df'] is not None else None
        self.split_all_students = state['split_all_students']
        self.student_mapping = state['student_mapping'].copy() if state['student_mapping'] else {}
        return True, "Đã hoàn tác"

    def redo(self) -> Tuple[bool, str]:
        if self.current_index >= len(self.history) - 1:
            return False, "Không thể làm lại thêm"
        self.current_index += 1
        state = self.history[self.current_index]
        self.original_df = state['original_df'].copy()
        self.filtered_df = state['filtered_df'].copy() if state['filtered_df'] is not None else None
        self.split_all_students = state['split_all_students']
        self.student_mapping = state['student_mapping'].copy() if state['student_mapping'] else {}
        return True, "Đã làm lại"