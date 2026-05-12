# app/utils/legacy/filter_v2.py
# -*- coding: utf-8 -*-
"""
ErrorExtractor - trích xuất lỗi cá nhân, phân loại bằng Rule + GPT
Không chứa giao diện Tkinter.
"""

import os
import re
import json
import traceback
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from openai import OpenAI
except ImportError:
    OpenAI = None

class Config:
    MODEL_NAME = "gpt-4o-mini"
    APP_DATA_DIR = "app_data"
    API_KEY_FILE = "apikey.json"

def get_api_key():
    try:
        os.makedirs(Config.APP_DATA_DIR, exist_ok=True)
        key_path = os.path.join(Config.APP_DATA_DIR, Config.API_KEY_FILE)
        if os.path.exists(key_path):
            with open(key_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('api_key', '')
    except:
        pass
    return ''

def save_api_key(api_key):
    try:
        os.makedirs(Config.APP_DATA_DIR, exist_ok=True)
        key_path = os.path.join(Config.APP_DATA_DIR, Config.API_KEY_FILE)
        with open(key_path, 'w', encoding='utf-8') as f:
            json.dump({'api_key': api_key}, f)
        return True
    except:
        return False

class GPTViolationClassifier:
    def __init__(self, api_key, model=Config.MODEL_NAME):
        if OpenAI is None:
            raise ImportError("Thư viện openai chưa cài đặt")
        self.client = OpenAI(api_key=api_key)
        self.model = model

    def classify(self, text):
        system_msg = (
            "Bạn là công cụ phân loại nhận xét học sinh. "
            "Phân thành 4 loại: vi_pham, can_cai_thien, khac, tich_cuc.\n"
            "... (giữ nguyên system prompt)"
        )
        user_prompt = f'Nhận xét: "{text}"\nLoại:'
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": system_msg},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0,
                max_tokens=10
            )
            answer = response.choices[0].message.content.strip().lower()
            if answer not in ["vi_pham", "can_cai_thien", "khac", "tich_cuc"]:
                return "khac"
            return answer
        except Exception:
            return "khac"

class ErrorExtractor:
    def __init__(self):
        self.input_df = None
        self.individual_errors = []
        self.classifier = None
        self.stats = {'rule': 0, 'gpt': 0}

        self.violation_patterns = [
            r'chưa\s+hoàn\s+thành|chưa\s+làm|không\s+làm|chưa\s+xong|không\s+hoàn\s+thành',
            r'không\s+mang\s+vở|quên\s+vở|không\s+có\s+vở|không\s+mang\s+sách|không\s+mang\s+võ\s+phục|sai\s+đồng\s+phục',
            # ... (toàn bộ patterns)
        ]

        self.improvement_patterns = [
            r'cần\s+(tập\s+trung|cố\s+gắng|chú\s+ý|điều\s+chỉnh|bổ\s+sung|xem\s+lại)',
            # ...
        ]

        self.positive_patterns = [
            r'tuyên\s+dương|khen|tốt|giỏi|xuất\s+sắc|tiến\s+bộ|tích\s+cực',
            # ...
        ]

        self.other_patterns = [
            r'chép\s+phạt|P;|^P$|^KP$|xuống\s+y\s+tế|y\s+tế|xin\s+xuống',
        ]

    def set_api_key(self, api_key):
        if api_key and OpenAI is not None:
            self.classifier = GPTViolationClassifier(api_key)
        else:
            self.classifier = None

    def load_input_file(self, file_path):
        try:
            self.input_df = pd.read_excel(file_path, dtype=object, engine='openpyxl')
            self.input_df.columns = self.input_df.columns.str.strip()
            return True, f"Đã tải {len(self.input_df)} dòng"
        except Exception as e:
            return False, str(e)

    def _classify_text(self, text, use_gpt=True):
        text_lower = text.lower()
        for pat in self.violation_patterns:
            if re.search(pat, text_lower):
                return "vi_pham"
        for pat in self.improvement_patterns:
            if re.search(pat, text_lower):
                return "can_cai_thien"
        for pat in self.positive_patterns:
            if re.search(pat, text_lower):
                return "tich_cuc"
        for pat in self.other_patterns:
            if re.search(pat, text_lower):
                return "khac"
        if use_gpt and self.classifier is not None:
            return self.classifier.classify(text)
        return "khac"

    def extract_errors(self, progress_callback=None, use_gpt=True):
        if self.input_df is None:
            return False, "Chưa tải dữ liệu"
        self.individual_errors = []
        self.stats = {'rule': 0, 'gpt': 0}
        total_rows = len(self.input_df)
        for idx, row in self.input_df.iterrows():
            if progress_callback and idx % 20 == 0:
                progress_callback(f"Đang xử lý dòng {idx+1}/{total_rows}")
            student_id = row.get('học sinh', '')
            evaluation = row.get('đánh giá', '')
            if not (pd.notna(student_id) and str(student_id).strip()):
                continue
            if not (pd.notna(evaluation) and str(evaluation).strip()):
                continue
            eval_str = str(evaluation).strip()
            if eval_str.upper() in ['P', 'KP']:
                continue
            loai = self._classify_text(eval_str, use_gpt=False)
            if loai in ["vi_pham", "can_cai_thien", "tich_cuc"]:
                self.stats['rule'] += 1
            elif loai == "khac" and any(re.search(p, eval_str.lower()) for p in self.other_patterns):
                self.stats['rule'] += 1
            else:
                if use_gpt and self.classifier is not None:
                    loai = self.classifier.classify(eval_str)
                    self.stats['gpt'] += 1
                else:
                    loai = "khac"
                    self.stats['rule'] += 1
            self.individual_errors.append({
                'Ngày tháng': row.get('Ngày', ''),
                'Tiết': row.get('Tiết', ''),
                'Lớp': row.get('Lớp', ''),
                'Môn học': row.get('Môn học', ''),
                'Nguồn ghi nhận': row.get('Người dạy', ''),
                'Họ và tên': row.get('Họ tên học sinh', student_id),
                'Nội dung vi phạm': eval_str,
                'Loại': loai
            })
        self._sort_individual_errors()
        return True, f"Đã trích xuất {len(self.individual_errors)} nhận xét"

    def _sort_individual_errors(self):
        # sắp xếp theo lớp, ngày
        class_order = {'6A1':1, '6A2':2, '6A3':3, '7A1':4}
        def sort_key(err):
            cls = err['Lớp']
            order = class_order.get(cls, 999)
            date_val = err['Ngày tháng']
            if pd.isna(date_val) or str(date_val).strip()=='':
                return (order, 0)
            try:
                date_str = str(date_val)
                date_obj = datetime.strptime(date_str, '%d-%m-%Y')
            except:
                return (order, 0)
            return (order, -date_obj.timestamp())
        self.individual_errors.sort(key=sort_key)

    def generate_individual_report(self, template_path, output_path):
        import shutil
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Không tìm thấy file mẫu: {template_path}")
        shutil.copy(template_path, output_path)
        wb = load_workbook(output_path)
        if "Tong hop" not in wb.sheetnames:
            raise ValueError("File mẫu không có sheet 'Tong hop'")
        template_sheet = wb["Tong hop"]
        ws_tonghop = wb["Tong hop"]
        for row in ws_tonghop.iter_rows(min_row=2, max_row=ws_tonghop.max_row):
            for cell in row:
                cell.value = None
        vi_pham_data = [e for e in self.individual_errors if e['Loại']=="vi_pham"]
        for i, err in enumerate(vi_pham_data):
            self._write_row_with_style(ws_tonghop, 2+i, err, template_sheet)

        # Xử lý các sheet bổ sung (Cần cải thiện, Tích cực, Khác)
        target_idx = wb.sheetnames.index("Lớp 7A1") if "Lớp 7A1" in wb.sheetnames else len(wb.sheetnames)-1
        for sheet_name, loai_code in [("Cần cải thiện","can_cai_thien"), ("Tích cực","tich_cuc"), ("Khác","khac")]:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(title=sheet_name, index=target_idx+1)
            # Copy header style
            for col in range(1, template_sheet.max_column+1):
                src = template_sheet.cell(row=1, column=col)
                dst = ws.cell(row=1, column=col)
                self._copy_style(src, dst)
                dst.value = src.value
            for col in range(1, template_sheet.max_column+1):
                letter = get_column_letter(col)
                if template_sheet.column_dimensions[letter].width:
                    ws.column_dimensions[letter].width = template_sheet.column_dimensions[letter].width
            filtered = [e for e in self.individual_errors if e['Loại']==loai_code]
            for i, err in enumerate(filtered):
                self._write_row_with_style(ws, 2+i, err, template_sheet)
            target_idx += 1
        wb.save(output_path)

    def _write_row_with_style(self, ws, row_idx, err, template_sheet):
        cols = [
            err['Ngày tháng'], err['Tiết'], err['Lớp'], err['Môn học'],
            err['Nguồn ghi nhận'], err['Họ và tên'], err['Nội dung vi phạm'], "", ""
        ]
        for col_idx, value in enumerate(cols, start=1):
            src = template_sheet.cell(row=2, column=col_idx)
            dst = ws.cell(row=row_idx, column=col_idx, value=value)
            self._copy_style(src, dst)

    def _copy_style(self, src, dst):
        if src.has_style:
            dst.font = src.font.copy()
            dst.border = src.border.copy()
            dst.fill = src.fill.copy()
            dst.number_format = src.number_format
            dst.alignment = src.alignment.copy()