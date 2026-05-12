# app/utils/legacy/xeploai.py
# -*- coding: utf-8 -*-
"""
TÍNH ĐIỂM RÈN LUYỆN – 4 chế độ xếp loại
KHÔNG CÓ GIAO DIỆN TKINTER.
"""

import pandas as pd
import os
import re
from datetime import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# Hàm xếp loại (chế độ 1)
# ------------------------------------------------------------
def xep_loai(diem):
    if diem >= 90:
        return "Tốt"
    elif diem >= 70:
        return "Khá"
    elif diem >= 50:
        return "Đạt"
    else:
        return "Chưa Đạt"

# ------------------------------------------------------------
# Đọc sheet lớp (chế độ 1) – linh hoạt tìm dòng tiêu đề
# ------------------------------------------------------------
def doc_sheet_lop_che_do_1(file_path, sheet_name):
    try:
        df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        header_row = None
        for i, row in df_raw.iterrows():
            row_str = row.astype(str).str.lower().tolist()
            for cell in row_str:
                if 'mã hs' in cell or 'ma hs' in cell:
                    header_row = i
                    break
            if header_row is not None:
                break
        if header_row is None:
            return None
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
        df.columns = df.columns.str.strip()
        required = ['Mã HS', 'Họ và tên', 'Giới tính', 'Tổng điểm']
        for col in required:
            if col not in df.columns:
                return None
        return df
    except Exception:
        return None

# ------------------------------------------------------------
# Chế độ 1: Tính điểm trung bình từng học sinh
# ------------------------------------------------------------
def xu_ly_che_do_1(file_paths):
    data = {}
    for file_path in file_paths:
        if file_path is None:
            continue
        xl = pd.ExcelFile(file_path)
        for sheet in xl.sheet_names:
            if sheet == "Tong hop":
                continue
            df = doc_sheet_lop_che_do_1(file_path, sheet)
            if df is None:
                continue
            for _, row in df.iterrows():
                ma_hs = row['Mã HS']
                if pd.isna(ma_hs):
                    continue
                ho_ten = row['Họ và tên']
                gioi_tinh = row['Giới tính']
                tong_diem = row['Tổng điểm']
                if pd.isna(tong_diem):
                    continue
                try:
                    diem = float(tong_diem)
                except:
                    continue
                key = (sheet, ma_hs)
                if key not in data:
                    data[key] = {
                        'ho_ten': ho_ten,
                        'gioi_tinh': gioi_tinh,
                        'diem_list': []
                    }
                data[key]['diem_list'].append(diem)

    if not data:
        return None

    lop_dict = {}
    for (lop, ma_hs), info in data.items():
        diem_tb = sum(info['diem_list']) / len(info['diem_list'])
        xl = xep_loai(diem_tb)
        if lop not in lop_dict:
            lop_dict[lop] = []
        lop_dict[lop].append({
            'Mã HS': ma_hs,
            'Họ và tên': info['ho_ten'],
            'Giới tính': info['gioi_tinh'],
            'Điểm trung bình': round(diem_tb, 2),
            'Xếp loại': xl
        })

    # Sắp xếp theo file đầu tiên
    first_file = next((f for f in file_paths if f is not None), None)
    if first_file:
        try:
            xl_first = pd.ExcelFile(first_file)
            for sheet in xl_first.sheet_names:
                if sheet in lop_dict:
                    df_order = doc_sheet_lop_che_do_1(first_file, sheet)
                    if df_order is not None:
                        order = {row['Mã HS']: idx for idx, row in df_order.iterrows()}
                        lop_dict[sheet].sort(key=lambda x: order.get(x['Mã HS'], 999999))
        except:
            pass
    return lop_dict

# ------------------------------------------------------------
# Chế độ 2: Tính trung bình cộng điểm của học sinh (lỗi cá nhân)
# ------------------------------------------------------------
def tinh_trung_binh_lop(file_path, sheet_name):
    try:
        df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        header_row = None
        for i, row in df_raw.iterrows():
            row_str = row.astype(str).str.lower().tolist()
            if any('mã hs' in cell for cell in row_str):
                header_row = i
                break
        if header_row is None:
            return None
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
        df.columns = df.columns.str.strip()
        if 'Mã HS' not in df.columns or 'Tổng điểm' not in df.columns:
            return None
        df_hs = df[df['Mã HS'].notna()].copy()
        if df_hs.empty:
            return None
        diem_series = pd.to_numeric(df_hs['Tổng điểm'], errors='coerce')
        diem_series = diem_series.dropna()
        if len(diem_series) == 0:
            return None
        return round(diem_series.mean(), 2)
    except Exception as e:
        print(f"Lỗi tính trung bình lớp {sheet_name}: {e}")
        return None

def xu_ly_che_do_2(file_paths):
    all_classes = set()
    week_data = []
    for idx, file_path in enumerate(file_paths, start=1):
        if file_path is None:
            week_data.append({})
            continue
        xl = pd.ExcelFile(file_path)
        week_dict = {}
        for sheet in xl.sheet_names:
            if sheet == "Tong hop":
                continue
            tb = tinh_trung_binh_lop(file_path, sheet)
            if tb is not None:
                week_dict[sheet] = tb
                all_classes.add(sheet)
        week_data.append(week_dict)

    if not any(week_data):
        return None

    df_result = pd.DataFrame(week_data)
    df_result = df_result.reindex(columns=sorted(all_classes))
    df_result.insert(0, 'Tuần', [f'Tuần {i+1}' for i in range(len(week_data))])

    numeric_cols = [col for col in df_result.columns if col != 'Tuần']
    avg_row = {}
    for col in numeric_cols:
        avg_row[col] = df_result[col].mean()
    avg_df = pd.DataFrame([avg_row])
    avg_df.insert(0, 'Tuần', 'TỔNG')
    df_result = pd.concat([df_result, avg_df], ignore_index=True)

    for col in numeric_cols:
        df_result[col] = df_result[col].round(2)
    return df_result

# ------------------------------------------------------------
# Chế độ 3: Đọc file lỗi tập thể, tính tổng điểm trừ theo lớp từng tuần
# ------------------------------------------------------------
def doc_sheet_loi_tap_the(file_path, sheet_name):
    try:
        df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        header_row = None
        for i, row in df_raw.iterrows():
            row_str = row.astype(str).str.lower().tolist()
            has_lop = any('lớp' in cell or 'lop' in cell for cell in row_str)
            has_diemtru = any('điểm trừ' in cell or 'diem tru' in cell for cell in row_str)
            if has_lop and has_diemtru:
                header_row = i
                break
        if header_row is None:
            return None
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
        df.columns = df.columns.str.strip()
        col_lop = None
        col_diem = None
        for col in df.columns:
            col_low = col.lower()
            if 'lớp' in col_low or 'lop' in col_low:
                col_lop = col
            if 'điểm trừ' in col_low or 'diem tru' in col_low:
                col_diem = col
        if col_lop is None or col_diem is None:
            return None
        df_valid = df[[col_lop, col_diem]].dropna(how='any')
        df_valid[col_diem] = pd.to_numeric(df_valid[col_diem], errors='coerce')
        df_valid = df_valid.dropna(subset=[col_diem])
        if df_valid.empty:
            return None
        return df_valid[[col_lop, col_diem]]
    except Exception:
        return None

def xu_ly_che_do_3(file_paths):
    default_classes = ['6A1', '6A2', '6A3', '7A1']
    week_data = []
    for idx, file_path in enumerate(file_paths, start=1):
        if file_path is None:
            week_data.append({})
            continue
        xl = pd.ExcelFile(file_path)
        week_dict = {}
        pattern = re.compile(r'^[Tt]u[âa]?n\s*\d+')
        for sheet in xl.sheet_names:
            if pattern.match(sheet):
                df = doc_sheet_loi_tap_the(file_path, sheet)
                if df is not None:
                    col_lop, col_diem = df.columns
                    for _, row in df.iterrows():
                        lop = str(row[col_lop]).strip()
                        diem = float(row[col_diem])
                        week_dict[lop] = week_dict.get(lop, 0) + diem
        week_data.append(week_dict)

    rows = []
    for i, wdict in enumerate(week_data, start=1):
        row = [f'Tuần {i}']
        for cls in default_classes:
            row.append(wdict.get(cls))
        rows.append(row)

    df_result = pd.DataFrame(rows, columns=['TUẦN'] + default_classes)

    total_row = ['TỔNG']
    for cls in default_classes:
        col_sum = df_result[cls].sum(skipna=True)
        total_row.append(0 if pd.isna(col_sum) else col_sum)
    df_result.loc[len(df_result)] = total_row

    last_row = df_result.iloc[-1]
    diem_tap_the = ['ĐIỂM TẬP THỂ']
    for cls in default_classes:
        diem_tap_the.append(100 - last_row[cls])
    df_result.loc[len(df_result)] = diem_tap_the

    for cls in default_classes:
        df_result[cls] = df_result[cls].apply(lambda x: int(x) if pd.notna(x) else x)
    return df_result

# ------------------------------------------------------------
# Chế độ 4: Lọc danh sách vi phạm cá nhân từ nhiều file (sheet "Tong hop")
# ------------------------------------------------------------
def xu_ly_che_do_4(file_paths):
    all_data = []
    error_messages = []
    required_patterns = {
        'ngày tháng': r'ngày.*tháng|ngay.*thang',
        'lớp': r'\b(lớp|lop)\b',
        'môn học': r'môn|mon',
        'nguồn ghi nhận': r'nguồn|nguon',
        'họ và tên': r'(họ.*tên|ho.*ten|họ tên|ho ten)',
        'nội dung vi phạm': r'nội dung|noi dung',
        'lỗi cấp độ': r'lỗi.*cấp|loi.*cap'
    }
    for file_path in file_paths:
        if not os.path.exists(file_path):
            error_messages.append(f"File không tồn tại: {file_path}")
            continue
        try:
            xl = pd.ExcelFile(file_path)
            sheet_name = None
            for sn in xl.sheet_names:
                if sn.lower().strip() == "tong hop":
                    sheet_name = sn
                    break
            if sheet_name is None:
                error_messages.append(f"File {os.path.basename(file_path)}: Không có sheet 'Tong hop'")
                continue
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            df.columns = df.columns.str.strip()
            found_cols = {}
            for col in df.columns:
                col_lower = col.lower()
                for key, pattern in required_patterns.items():
                    if key not in found_cols and re.search(pattern, col_lower):
                        found_cols[key] = col
                        break
            if 'ngày tháng' not in found_cols:
                error_messages.append(f"File {os.path.basename(file_path)}: Thiếu cột ngày tháng (các cột: {list(df.columns)})")
                continue
            selected_cols = [found_cols[k] for k in found_cols if k in found_cols]
            if not selected_cols:
                error_messages.append(f"File {os.path.basename(file_path)}: Không tìm thấy cột nào")
                continue
            df_sub = df[selected_cols].copy()
            rename_dict = {found_cols[k]: k for k in found_cols}
            df_sub.rename(columns=rename_dict, inplace=True)
            try:
                df_sub['ngày tháng'] = pd.to_datetime(df_sub['ngày tháng'], errors='coerce', dayfirst=True)
            except:
                df_sub['ngày tháng'] = pd.to_datetime(df_sub['ngày tháng'], errors='coerce')
            df_sub = df_sub.dropna(subset=['ngày tháng'])
            if df_sub.empty:
                error_messages.append(f"File {os.path.basename(file_path)}: Không có ngày hợp lệ")
                continue
            if 'lớp' in df_sub.columns:
                df_sub['lớp'] = df_sub['lớp'].astype(str).str.strip()
            all_data.append(df_sub)
        except Exception as e:
            error_messages.append(f"File {os.path.basename(file_path)}: Lỗi - {str(e)}")
    if not all_data:
        return None, error_messages
    df_final = pd.concat(all_data, ignore_index=True)
    df_final = df_final.sort_values(by='ngày tháng', ascending=True)
    df_final['ngày tháng'] = df_final['ngày tháng'].dt.strftime('%d/%m/%Y')
    ordered_cols = ['ngày tháng', 'lớp', 'môn học', 'nguồn ghi nhận', 'họ và tên', 'nội dung vi phạm', 'lỗi cấp độ']
    final_cols = [col for col in ordered_cols if col in df_final.columns]
    df_final = df_final[final_cols]
    return df_final, error_messages

# ------------------------------------------------------------
# Lưu kết quả (hỗ trợ cả 4 chế độ)
# ------------------------------------------------------------
def save_results(result, output_file, mode):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        if mode == 1:
            for lop, students in result.items():
                df_out = pd.DataFrame(students)
                df_out.insert(0, 'STT', range(1, len(df_out)+1))
                df_out.to_excel(writer, sheet_name=lop, index=False)
        elif mode == 2:
            result.to_excel(writer, sheet_name='Tong_hop', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Tong_hop']
            worksheet.cell(row=1, column=1, value="TUẦN")
            header_font = Font(bold=True)
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for row in range(2, worksheet.max_row + 1):
                for col in range(2, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='center')
            total_row = None
            for row in range(2, worksheet.max_row + 1):
                if worksheet.cell(row=row, column=1).value == "TỔNG":
                    total_row = row
                    break
            if total_row:
                bold_font = Font(bold=True)
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=total_row, column=col)
                    cell.font = bold_font
                    if col >= 2 and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal='center')
            for col in range(1, worksheet.max_column + 1):
                worksheet.column_dimensions[get_column_letter(col)].auto_size = True
        elif mode == 3:
            result.to_excel(writer, sheet_name='Tong_hop', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Tong_hop']
            worksheet.cell(row=1, column=1, value="TUẦN")
            header_font = Font(bold=True)
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for row in range(2, worksheet.max_row + 1):
                for col in range(2, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '0'
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        cell.alignment = Alignment(horizontal='center')
            bold_font = Font(bold=True)
            for row in range(2, worksheet.max_row + 1):
                val = worksheet.cell(row=row, column=1).value
                if val in ("TỔNG", "ĐIỂM TẬP THỂ"):
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = bold_font
                        if col >= 2 and isinstance(cell.value, (int, float)):
                            cell.number_format = '0'
                        cell.alignment = Alignment(horizontal='center')
            for col in range(1, worksheet.max_column + 1):
                worksheet.column_dimensions[get_column_letter(col)].auto_size = True
        else:  # mode == 4
            result.to_excel(writer, sheet_name='Danh_sach_vi_pham', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Danh_sach_vi_pham']
            header_font = Font(bold=True)
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in range(1, worksheet.max_column + 1):
                worksheet.column_dimensions[get_column_letter(col)].auto_size = True